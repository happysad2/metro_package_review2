from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


REQUIRED_SHEETS = [
    "Standard governance",
    "Location Specification",
    "Location List - To Be Populated",
]

ASSET_LIST_SHEET = "Asset List - To Be Populated"

REQUIRED_HEADER_CELLS = {
    "H6": "Asset Description",
    "T6": "GPS Coordinates",
    "AA6": "Coordinate Datum",
    "J6": "Uniclass Code",
}

COL_ASSET_CODE = 5   # E
COL_DATUM = 27       # AA (fallback bound for export width)
HEADER_ROW = 6
DATA_START_ROW = 7

HEADER_ASSET_DESCRIPTION = "Asset Description"
HEADER_GPS = "GPS Coordinates"
HEADER_DATUM = "Coordinate Datum"
HEADER_UNICLASS = "Uniclass Code"


@dataclass
class ValidationResult:
    source_file: Path
    passed: bool = True
    template_checks_passed: bool = True
    has_value_checks_passed: bool = True
    template_comments: List[str] = field(default_factory=list)
    has_value_comments: List[str] = field(default_factory=list)
    template_details: List[str] = field(default_factory=list)
    has_value_details: List[str] = field(default_factory=list)
    detailed_log: List[str] = field(default_factory=list)
    exported_csv: Path | None = None

    def fail(self, section: str, concise_comment: str, detailed_comment: str) -> None:
        self.passed = False
        if section == "template":
            self.template_checks_passed = False
            if concise_comment not in self.template_comments:
                self.template_comments.append(concise_comment)
            self.template_details.append(detailed_comment)
        else:
            self.has_value_checks_passed = False
            if concise_comment not in self.has_value_comments:
                self.has_value_comments.append(concise_comment)
            self.has_value_details.append(detailed_comment)
        self.detailed_log.append(detailed_comment)

    def log(self, message: str) -> None:
        self.detailed_log.append(message)


def normalize_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_system_asset(asset_description: str) -> bool:
    lowered = asset_description.lower()
    return "system" in lowered or "systems" in lowered


def get_last_data_bounds(ws: Worksheet) -> Tuple[int, int]:
    last_row = 0
    last_col = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_has_data = False
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                row_has_data = True
                if cell.column > last_col:
                    last_col = cell.column
        if row_has_data:
            last_row = row[0].row

    if last_row == 0:
        return HEADER_ROW, COL_DATUM

    if last_col == 0:
        last_col = COL_DATUM

    return max(last_row, HEADER_ROW), max(last_col, COL_DATUM)


def export_sheet_to_csv(ws: Worksheet, output_csv: Path, last_row: int, last_col: int) -> None:
    with output_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col, values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def find_header_column(ws: Worksheet, header_name: str, max_col: int) -> int | None:
    for col_num in range(1, max_col + 1):
        header_value = normalize_value(ws.cell(row=HEADER_ROW, column=col_num).value)
        if header_value == header_name:
            return col_num
    return None


def find_required_columns(result: ValidationResult, ws: Worksheet, max_col: int) -> dict[str, int]:
    required_headers = [
        HEADER_ASSET_DESCRIPTION,
        HEADER_GPS,
        HEADER_DATUM,
        HEADER_UNICLASS,
    ]

    header_columns: dict[str, int] = {}
    missing_headers: List[str] = []

    for header in required_headers:
        col_num = find_header_column(ws, header, max_col)
        if col_num is None:
            missing_headers.append(header)
        else:
            header_columns[header] = col_num

    if missing_headers:
        result.fail(
            "template",
            "not in Sydney Metro Template, cells not aligning.",
            f"Missing required header(s) on row {HEADER_ROW}: {', '.join(missing_headers)}",
        )

    return header_columns


def check_template_structure(result: ValidationResult, workbook_sheet_names: List[str], ws_asset: Worksheet | None) -> None:
    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook_sheet_names]
    if missing_sheets:
        result.fail(
            "template",
            "not in Sydney Metro Template, cells not aligning.",
            f"Missing required sheet(s): {', '.join(missing_sheets)}",
        )

    if ws_asset is None:
        result.fail(
            "template",
            "not in Sydney Metro Template, cells not aligning.",
            f"Missing required sheet for header checks/export: {ASSET_LIST_SHEET}",
        )
        return

    for cell_ref, expected_value in REQUIRED_HEADER_CELLS.items():
        actual = normalize_value(ws_asset[cell_ref].value)
        if actual != expected_value:
            result.fail(
                "template",
                "not in Sydney Metro Template, cells not aligning.",
                f"Header mismatch at {cell_ref}: expected '{expected_value}', found '{actual or '[blank]'}'",
            )


def validate_asset_rows(
    result: ValidationResult,
    ws_asset: Worksheet,
    last_row: int,
    col_asset_desc: int,
    col_gps: int,
    col_datum: int,
    col_uniclass: int,
) -> None:
    seen_asset_codes: dict[str, int] = {}

    for row_num in range(DATA_START_ROW, last_row + 1):
        asset_code = normalize_value(ws_asset.cell(row=row_num, column=COL_ASSET_CODE).value)
        asset_desc = normalize_value(ws_asset.cell(row=row_num, column=col_asset_desc).value)
        uniclass = normalize_value(ws_asset.cell(row=row_num, column=col_uniclass).value)
        gps = normalize_value(ws_asset.cell(row=row_num, column=col_gps).value)
        datum = normalize_value(ws_asset.cell(row=row_num, column=col_datum).value)

        has_any_asset_data = any([asset_code, asset_desc, uniclass, gps, datum])
        if not has_any_asset_data:
            continue

        if not asset_desc:
            result.fail(
                "has_value",
                "one or more required fields are missing.",
                f"Row {row_num}: missing Asset Description",
            )

        if not asset_code:
            result.fail(
                "has_value",
                "asset code issues found.",
                f"Row {row_num}: missing Asset Code (column E)",
            )
        else:
            if asset_code in seen_asset_codes:
                result.fail(
                    "has_value",
                    "asset code issues found.",
                    f"Row {row_num}: duplicate Asset Code '{asset_code}' also appears on row {seen_asset_codes[asset_code]}",
                )
            else:
                seen_asset_codes[asset_code] = row_num

        if not is_system_asset(asset_desc):
            if not gps:
                result.fail(
                    "has_value",
                    "non-system asset location fields are incomplete.",
                    f"Row {row_num}: non-system asset missing GPS Coordinates",
                )
            if not datum:
                result.fail(
                    "has_value",
                    "non-system asset location fields are incomplete.",
                    f"Row {row_num}: non-system asset missing Coordinate Datum",
                )
            if not uniclass:
                result.fail(
                    "has_value",
                    "non-system asset location fields are incomplete.",
                    f"Row {row_num}: non-system asset missing Uniclass Code",
                )


def validate_workbook(xlsx_file: Path, output_dir: Path) -> ValidationResult:
    result = ValidationResult(source_file=xlsx_file)
    result.log(f"Validating file: {xlsx_file.name}")

    workbook = load_workbook(xlsx_file, data_only=True)
    sheet_names = workbook.sheetnames
    ws_asset = workbook[ASSET_LIST_SHEET] if ASSET_LIST_SHEET in sheet_names else None

    check_template_structure(result, sheet_names, ws_asset)

    if ws_asset is not None:
        last_row, last_col = get_last_data_bounds(ws_asset)
        csv_name = f"{xlsx_file.stem}_{ASSET_LIST_SHEET.replace(' ', '_')}.csv"
        csv_path = output_dir / csv_name
        export_sheet_to_csv(ws_asset, csv_path, last_row, last_col)
        result.exported_csv = csv_path
        result.log(f"Exported sheet '{ASSET_LIST_SHEET}' to CSV: {csv_path.name}")

        header_columns = find_required_columns(result, ws_asset, last_col)
        if all(
            header in header_columns
            for header in [HEADER_ASSET_DESCRIPTION, HEADER_GPS, HEADER_DATUM, HEADER_UNICLASS]
        ):
            validate_asset_rows(
                result,
                ws_asset,
                last_row,
                header_columns[HEADER_ASSET_DESCRIPTION],
                header_columns[HEADER_GPS],
                header_columns[HEADER_DATUM],
                header_columns[HEADER_UNICLASS],
            )

    if result.passed:
        result.template_comments.append("right template checks passed.")
        result.has_value_comments.append("has value checks passed.")
        result.log("Validation status: PASS")
    else:
        result.log("Validation status: FAIL")

    return result


def write_outputs(result: ValidationResult, output_dir: Path) -> Tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{result.source_file.stem}_{timestamp}"

    concise_file = output_dir / f"{base_name}_result.txt"
    log_file = output_dir / f"{base_name}_log.txt"

    status = "PASS" if result.passed else "FAIL"

    concise_lines = [
        f"File: {result.source_file.name}",
        f"Result: {status}",
        "Right template checks:",
        f"- {'PASS' if result.template_checks_passed else 'FAIL'}",
    ]

    if result.template_checks_passed:
        concise_lines.append("- right template checks passed.")
    else:
        concise_lines.append("- Submission template structure has been modified. Sydney Metro Template, cells not aligning.")
        concise_lines.extend(f"- {detail}" for detail in result.template_details)

    concise_lines.append("Has value checks:")
    concise_lines.append(f"- {'PASS' if result.has_value_checks_passed else 'FAIL'}")

    if result.has_value_checks_passed:
        concise_lines.append("- has value checks passed.")
    else:
        concise_lines.append("- Values are missing from your submission. This includes but is not limited to:")
        concise_lines.extend(f"- {detail}" for detail in result.has_value_details)

    if result.exported_csv:
        concise_lines.append(f"- exported CSV: {result.exported_csv.name}")

    log_lines = [
        f"File: {result.source_file}",
        f"Result: {status}",
        f"Timestamp: {datetime.now().isoformat(timespec='seconds')}",
        f"Right template checks: {'PASS' if result.template_checks_passed else 'FAIL'}",
        f"Has value checks: {'PASS' if result.has_value_checks_passed else 'FAIL'}",
    ]
    if result.exported_csv:
        log_lines.append(f"Exported CSV: {result.exported_csv}")

    log_lines.append("Details:")
    log_lines.extend(f"- {entry}" for entry in result.detailed_log)

    concise_file.write_text("\n".join(concise_lines) + "\n", encoding="utf-8")
    log_file.write_text("\n".join(log_lines) + "\n", encoding="utf-8")

    return concise_file, log_file


def find_input_files(input_dir: Path) -> List[Path]:
    files = sorted(list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xlsm")))
    return [file for file in files if file.is_file()]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate Sydney Metro asset register template and content checks."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("in"),
        help="Folder containing input Excel file(s). Default: in",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("out"),
        help="Folder to write CSV, concise results, and logs. Default: out",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_dir = args.input_dir
    output_dir = args.output_dir

    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_dir.exists():
        print(f"Input directory not found: {input_dir}")
        return 1

    input_files = find_input_files(input_dir)
    if not input_files:
        print(f"No Excel files found in: {input_dir}")
        return 1

    overall_pass = True
    for input_file in input_files:
        result = validate_workbook(input_file, output_dir)
        concise_file, log_file = write_outputs(result, output_dir)
        status = "PASS" if result.passed else "FAIL"
        print(f"{input_file.name}: {status}")
        print(f"  concise output: {concise_file}")
        print(f"  detailed log:   {log_file}")
        if result.exported_csv:
            print(f"  exported csv:   {result.exported_csv}")
        if not result.passed:
            overall_pass = False

    return 0 if overall_pass else 2


if __name__ == "__main__":
    raise SystemExit(main())
