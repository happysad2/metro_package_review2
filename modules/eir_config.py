"""
EIR Configuration Module
========================
Discovers available EIR versions from the EIRs/ folder, locates their
schema files (BIM Schema, Asset Register, CAD Schema), and parses the
BIM Configuration Spec sheet to extract required properties per IFC
property set.

Drop a new EIR version folder (e.g. ``v8.3/``) into the EIRs directory
with its schema Excel files and this module will auto-discover it.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from modules import APP_ROOT

# ---------------------------------------------------------------------------
# EIR root — beside the application
# ---------------------------------------------------------------------------

import sys as _sys

if getattr(_sys, "frozen", False):
    _USER_ROOT = Path(_sys.executable).parent.resolve()
else:
    _USER_ROOT = APP_ROOT

EIR_ROOT = _USER_ROOT / "EIRs"

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class EIRVersion:
    """Represents a single discoverable EIR version."""
    version_key: str                     # sort key, e.g. "8.2", "6.01"
    display_name: str                    # e.g. "EIR v8.2"
    folder: Path
    bim_schema_path: Optional[Path] = None
    asset_register_path: Optional[Path] = None
    cad_schema_path: Optional[Path] = None
    has_schemas: bool = False


@dataclass
class BIMField:
    """One field parsed from the BIM Configuration Spec sheet."""
    field_no: int
    field_name: str
    prop_set: str                        # e.g. SM_Project, SM_Asset, SM_Location
    attr_name: str
    attr_level: str                      # Model | Object | ""
    mandatory: str
    phase_req: Dict[str, str] = field(default_factory=dict)
    fmt: str = ""


@dataclass
class BIMSchemaConfig:
    """Parsed BIM Schema for a given EIR version."""
    version: str
    fields: List[BIMField] = field(default_factory=list)

    # Derived: IFC property-set → list of attribute names
    ifc_property_sets: Dict[str, List[str]] = field(default_factory=dict)

    def build_property_sets(self) -> None:
        """Group attributes by IFC Property Set."""
        self.ifc_property_sets.clear()
        for f in self.fields:
            ps = f.prop_set.strip()
            if ps:
                self.ifc_property_sets.setdefault(ps, []).append(f.attr_name)


# ---------------------------------------------------------------------------
# Version discovery
# ---------------------------------------------------------------------------

_VER_RE = re.compile(r"^v?(\d+(?:\.\d+)*)$", re.I)


def _version_sort_key(name: str) -> tuple:
    """Return a tuple for natural numeric sorting of version strings."""
    m = _VER_RE.match(name)
    if not m:
        return (999,)
    return tuple(int(p) for p in m.group(1).split("."))


def _find_schema_file(folder: Path, *patterns: str) -> Optional[Path]:
    """Walk schema sub-folders for a file matching any of the given patterns."""
    for child in sorted(folder.rglob("*")):
        if not child.is_file():
            continue
        if child.suffix.lower() not in (".xlsx", ".xls"):
            continue
        # Skip temp/superseded files
        if child.name.startswith("~$"):
            continue
        low = child.name.lower()
        rel = str(child.relative_to(folder)).lower()
        if "_superseded" in rel or "\\superseded" in rel:
            continue
        for pat in patterns:
            if pat.lower() in low:
                return child
    return None


def discover_versions(eir_root: Optional[Path] = None) -> List[EIRVersion]:
    """Scan the EIR root folder and return available versions, newest first."""
    root = eir_root or EIR_ROOT
    if not root.is_dir():
        return []

    versions: list[EIRVersion] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        m = _VER_RE.match(child.name)
        if not m:
            continue
        ver_str = m.group(1)
        display = f"EIR v{ver_str}"

        bim = _find_schema_file(child, "bim schema")
        ar  = _find_schema_file(child, "asset register")
        cad = _find_schema_file(child, "cad schema")

        eir = EIRVersion(
            version_key=ver_str,
            display_name=display,
            folder=child,
            bim_schema_path=bim,
            asset_register_path=ar,
            cad_schema_path=cad,
            has_schemas=bim is not None or ar is not None,
        )
        versions.append(eir)

    # Sort newest-first
    versions.sort(key=lambda v: _version_sort_key(v.version_key), reverse=True)
    return versions


# ---------------------------------------------------------------------------
# BIM Schema parsing (auto-detect column layout)
# ---------------------------------------------------------------------------

# Header patterns we look for (case-insensitive partial match)
_HDR_FIELD_NO  = {"field no", "field\nno"}
_HDR_FIELD_NAME = {"field name"}
_HDR_PROP_SET  = {"ifc property set", "property set"}
_HDR_ATTR_NAME = {"attribute name"}
_HDR_ATTR_LVL  = {"attribute level"}
_HDR_MANDATORY = {"mandatory"}
_HDR_FORMAT    = {"format", "format / constraint"}

# Phase header patterns → canonical names
_PHASE_PATTERNS = [
    ("business case",       "Business Case"),
    ("preliminary",         "Preliminary Design"),
    ("detailed design",     "Detailed Design"),
    ("procurement",         "Procurement"),
    ("test readiness",      "Test Readiness Review"),
    ("system verification", "System Verification Review"),
    ("operations",          "Operations and Maintenance"),
    # Older EIR names
    ("feasib",              "Business Case"),
    ("concept design",      "Preliminary Design"),
    ("afc",                 "Detailed Design"),
    ("construction",        "Procurement"),
    ("as-built",            "Test Readiness Review"),
    ("o&m",                 "Operations and Maintenance"),
    ("options",             "Business Case"),
]


def _match_header(val: str, patterns: set) -> bool:
    low = val.lower().strip()
    return any(p in low for p in patterns)


def _detect_columns(ws, scan_rows=(2, 3, 4)) -> dict:
    """Scan a few header rows to find column indices for key fields."""
    mapping = {}
    phase_cols: dict[str, int] = {}

    for row_num in scan_rows:
        for col in range(1, min(ws.max_column + 1, 30)):
            raw = ws.cell(row=row_num, column=col).value
            if raw is None:
                continue
            val = str(raw).strip()
            if not val:
                continue
            low = val.lower()

            if _match_header(val, _HDR_FIELD_NO) and "field_no" not in mapping:
                mapping["field_no"] = (row_num, col)
            elif _match_header(val, _HDR_FIELD_NAME) and "field_name" not in mapping:
                mapping["field_name"] = (row_num, col)
            elif _match_header(val, _HDR_PROP_SET) and "prop_set" not in mapping:
                mapping["prop_set"] = (row_num, col)
            elif _match_header(val, _HDR_ATTR_NAME) and "attr_name" not in mapping:
                mapping["attr_name"] = (row_num, col)
            elif _match_header(val, _HDR_ATTR_LVL) and "attr_level" not in mapping:
                mapping["attr_level"] = (row_num, col)
            elif _match_header(val, _HDR_MANDATORY) and "mandatory" not in mapping:
                mapping["mandatory"] = (row_num, col)
            elif _match_header(val, _HDR_FORMAT) and "fmt" not in mapping:
                mapping["fmt"] = (row_num, col)

            # Phase columns
            for pattern, canonical in _PHASE_PATTERNS:
                if pattern in low and canonical not in phase_cols:
                    phase_cols[canonical] = col
                    break

    # Work out the header row (most common row among detected columns)
    rows_found = [r for r, _ in mapping.values()]
    header_row = max(set(rows_found), key=rows_found.count) if rows_found else 3

    return {
        "header_row": header_row,
        "field_no": mapping.get("field_no", (header_row, 1))[1],
        "field_name": mapping.get("field_name", (header_row, 2))[1],
        "prop_set": mapping.get("prop_set", (header_row, 15))[1],
        "attr_name": mapping.get("attr_name", (header_row, 16))[1],
        "attr_level": mapping.get("attr_level", (0, 0))[1],  # 0 = not found
        "mandatory": mapping.get("mandatory", (0, 0))[1],
        "fmt": mapping.get("fmt", (0, 0))[1],
        "phase_cols": phase_cols,
    }


_BIM_SHEET_CANDIDATES = ["BIM Configuration Spec"]


def load_bim_schema(eir: EIRVersion) -> Optional[BIMSchemaConfig]:
    """Parse the BIM Schema for the given EIR version.

    Returns None if no schema file exists or parsing fails.
    """
    if not eir.bim_schema_path or not eir.bim_schema_path.exists():
        return None

    import openpyxl

    try:
        wb = openpyxl.load_workbook(str(eir.bim_schema_path), data_only=True)
    except Exception:
        return None

    ws = None
    for name in _BIM_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        return None

    cols = _detect_columns(ws)
    header_row = cols["header_row"]
    data_start = header_row + 1

    config = BIMSchemaConfig(version=eir.version_key)

    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
        fno_val = row[cols["field_no"] - 1].value
        if fno_val is None:
            continue
        try:
            fno = int(fno_val)
        except (ValueError, TypeError):
            continue

        field_name = str(row[cols["field_name"] - 1].value or "").strip()
        prop_set   = str(row[cols["prop_set"] - 1].value or "").strip()
        attr_name  = str(row[cols["attr_name"] - 1].value or "").strip()

        if not prop_set or not attr_name:
            continue

        attr_level = ""
        if cols["attr_level"]:
            attr_level = str(row[cols["attr_level"] - 1].value or "").strip()

        mandatory = ""
        if cols["mandatory"]:
            mandatory = str(row[cols["mandatory"] - 1].value or "").strip()

        fmt = ""
        if cols["fmt"]:
            fmt = str(row[cols["fmt"] - 1].value or "").strip()

        phase_req: dict[str, str] = {}
        for phase_name, phase_col in cols["phase_cols"].items():
            val = str(row[phase_col - 1].value or "").strip()
            phase_req[phase_name] = val if val else "Optional"

        config.fields.append(BIMField(
            field_no=fno,
            field_name=field_name,
            prop_set=prop_set,
            attr_name=attr_name,
            attr_level=attr_level,
            mandatory=mandatory,
            phase_req=phase_req,
            fmt=fmt,
        ))

    wb.close()
    config.build_property_sets()
    config._source_path = eir.bim_schema_path  # used by NWC checker
    return config
