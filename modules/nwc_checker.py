"""
NWC Checker Module
==================
Validates NWC (Navisworks Cache) files against the Sydney Metro BIM
Configuration Specification (SM-22-00340097).

Requirements: openpyxl, pywin32, Autodesk Navisworks Manage installed.
If Navisworks is not available the module reports that NWC checking is
unavailable rather than crashing.
"""

from __future__ import annotations

import csv
import os
from pathlib import Path

from modules import ModuleResult

# ---------------------------------------------------------------------------
# Schema constants
# ---------------------------------------------------------------------------

SCHEMA_XLSX_NAME = "SM-22-00340097 BIM Schema and Specification.xlsx"
SCHEMA_SHEET = "BIM Configuration Spec"
SCHEMA_HEADER_ROW = 3
SCHEMA_DATA_START = 4

PHASE_COLUMNS = {
    "Business Case": 6,
    "Preliminary Design": 7,
    "Detailed Design": 8,
    "Procurement": 9,
    "Test Readiness Review": 10,
    "System Verification Review": 11,
    "Operations and Maintenance": 12,
}

ATTR_NAME_MAP = {
    "tbProjectContractCode": "Project_Contract_Code",
    "SM_ABSID": "SM_SBS_ID",
    "tbAEODisciplineCode__SM_Asset": "TfNSW_DisciplineCode",
    "tbAEODisciplineDesc__SM_Asset": "TfNSW_SubDisciplineCode",
}

EMPTY_VALUES = {"", "N/A", "n/a", "NA", "na", "-", "--", "None", "none", "null"}

DEFAULT_PHASE = "Detailed Design"

# ---------------------------------------------------------------------------
# Schema loading
# ---------------------------------------------------------------------------

class SchemaField:
    __slots__ = ("field_no", "field_name", "mandatory", "attr_level",
                 "prop_set", "attr_name", "fmt", "phase_req", "nwc_prop_name")

    def __init__(self, field_no, field_name, mandatory, attr_level,
                 prop_set, attr_name, fmt, phase_req):
        self.field_no = field_no
        self.field_name = field_name
        self.mandatory = mandatory
        self.attr_level = attr_level
        self.prop_set = prop_set.strip()
        self.attr_name = attr_name.strip()
        self.fmt = fmt
        self.phase_req = phase_req
        compound = f"{self.attr_name}__{self.prop_set}"
        if compound in ATTR_NAME_MAP:
            self.nwc_prop_name = ATTR_NAME_MAP[compound]
        elif self.attr_name in ATTR_NAME_MAP:
            self.nwc_prop_name = ATTR_NAME_MAP[self.attr_name]
        else:
            self.nwc_prop_name = self.attr_name

    def is_mandatory_for_phase(self, phase: str) -> bool:
        return self.phase_req.get(phase, "Optional") == "Mandatory"

    def is_conditional_for_phase(self, phase: str) -> bool:
        return self.phase_req.get(phase, "Optional") == "Conditional"


def _find_schema_xlsx(app_root: Path) -> Path | None:
    from modules import REFERENCE_DOCS
    candidates = [
        REFERENCE_DOCS / SCHEMA_XLSX_NAME,
        app_root / "reference_docs" / SCHEMA_XLSX_NAME,
        app_root / SCHEMA_XLSX_NAME,
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _load_schema(xlsx_path: Path) -> list[SchemaField]:
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[SCHEMA_SHEET]
    fields: list[SchemaField] = []
    for row in ws.iter_rows(min_row=SCHEMA_DATA_START, max_row=ws.max_row, values_only=False):
        field_no = row[0].value
        if field_no is None:
            continue
        try:
            field_no = int(field_no)
        except (ValueError, TypeError):
            continue

        prop_set = str(row[14].value or "").strip()
        attr_name = str(row[15].value or "").strip()
        if not prop_set or not attr_name:
            continue

        phase_req = {}
        for pname, col_idx in PHASE_COLUMNS.items():
            val = str(row[col_idx].value or "").strip()
            phase_req[pname] = val if val else "Optional"

        fields.append(SchemaField(
            field_no,
            str(row[1].value or "").strip(),
            str(row[4].value or "").strip(),
            str(row[13].value or "").strip(),
            prop_set,
            attr_name,
            str(row[16].value or "").strip(),
            phase_req,
        ))
    wb.close()
    return fields


# ---------------------------------------------------------------------------
# NWC reading via Navisworks COM
# ---------------------------------------------------------------------------

class NwcObject:
    __slots__ = ("name", "class_name", "path", "properties", "depth")

    def __init__(self, name, class_name, path, depth):
        self.name = name
        self.class_name = class_name
        self.path = path
        self.depth = depth
        self.properties: dict[str, dict[str, str]] = {}


def _read_node_props(node) -> dict[str, dict[str, str]]:
    result = {}
    try:
        attrs = node.Attributes()
    except Exception:
        return result
    for j in range(1, attrs.Count + 1):
        attr = attrs.Item(j)
        try:
            cat = attr.ClassUserName
        except Exception:
            continue
        if not cat or not (cat.startswith("SM_") or cat == "Element"):
            continue
        props = {}
        try:
            pc = attr.Properties()
            for k in range(1, pc.Count + 1):
                p = pc.Item(k)
                try:
                    props[p.UserName] = str(p.value) if p.value is not None else ""
                except Exception:
                    props[getattr(p, "UserName", f"prop_{k}")] = ""
        except Exception:
            pass
        if props:
            result[cat] = props
    return result


def _traverse(node, objects: list[NwcObject], path_parts: list[str], depth: int):
    name = cls = ""
    try:
        name = node.UserName or ""
        cls = node.ClassName or ""
    except Exception:
        pass
    current = path_parts + ([name] if name else [])
    props = _read_node_props(node)
    if props:
        obj = NwcObject(name, cls, " > ".join(current), depth)
        obj.properties = props
        objects.append(obj)
    try:
        children = node.Children()
        for i in range(1, children.Count + 1):
            _traverse(children.Item(i), objects, current, depth + 1)
    except Exception:
        pass


def _read_nwc(nwc_path: str) -> list[NwcObject]:
    import win32com.client
    nwc_path = os.path.abspath(nwc_path)
    nw = win32com.client.Dispatch("Navisworks.Document")
    nw.OpenFile(nwc_path)
    partition = nw.State.CurrentPartition
    objects: list[NwcObject] = []
    _traverse(partition, objects, [], 0)
    return objects


# ---------------------------------------------------------------------------
# Checking
# ---------------------------------------------------------------------------

def _check_object(obj: NwcObject, fields: list[SchemaField], phase: str, result: ModuleResult, fname: str):
    for field in fields:
        if field.attr_level == "Model" and "SM_Project" not in obj.properties:
            continue
        if field.attr_level == "Object" and not any(
            ps in obj.properties for ps in ("SM_Asset", "SM_Location", "SM_Quantification")
        ):
            continue

        mandatory = field.is_mandatory_for_phase(phase)
        conditional = field.is_conditional_for_phase(phase)
        if not mandatory and not conditional:
            continue

        ps_data = obj.properties.get(field.prop_set, {})
        value = ps_data.get(field.nwc_prop_name)

        obj_label = obj.name or "(unnamed)"
        if value is None:
            if mandatory:
                result.add_finding(
                    fname, f"{field.field_name}", "FAIL",
                    f"Object '{obj_label}': mandatory attribute '{field.nwc_prop_name}' "
                    f"missing from '{field.prop_set}'"
                )
            elif conditional:
                result.add_finding(
                    fname, f"{field.field_name}", "WARNING",
                    f"Object '{obj_label}': conditional attribute '{field.nwc_prop_name}' not present"
                )
        elif value.strip() in EMPTY_VALUES:
            if mandatory:
                result.add_finding(
                    fname, f"{field.field_name}", "FAIL",
                    f"Object '{obj_label}': mandatory '{field.nwc_prop_name}' is empty ('{value}')"
                )
        else:
            if mandatory:
                result.add_finding(
                    fname, f"{field.field_name}", "PASS",
                    f"Object '{obj_label}': '{field.nwc_prop_name}' populated"
                )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run(input_folder: Path, output_folder: Path, log_callback=None,
        phase: str = DEFAULT_PHASE) -> ModuleResult:
    """
    Scan *input_folder* for .nwc files and validate each against the BIM schema.
    """
    result = ModuleResult(module_name="NWC Model")

    nwc_files = sorted(input_folder.glob("*.nwc"))
    nwc_files = [f for f in nwc_files if f.is_file()]

    if not nwc_files:
        result.build_granular_text()
        result.build_summary()
        if log_callback:
            log_callback("[NWC] No .nwc files found.")
        return result

    # Check dependencies
    try:
        import win32com.client  # noqa: F401
    except ImportError:
        for f in nwc_files:
            result.files_checked.append(f.name)
            result.add_finding(
                f.name, "Dependency check", "WARNING",
                "pywin32 is not installed — NWC checking requires pywin32 and Navisworks Manage."
            )
        result.build_granular_text()
        result.build_summary()
        if log_callback:
            log_callback("[NWC] pywin32 not available — skipping NWC checks.")
        return result

    # Find schema
    from modules import APP_ROOT as app_root
    schema_path = _find_schema_xlsx(app_root)
    if schema_path is None:
        for f in nwc_files:
            result.files_checked.append(f.name)
            result.add_finding(
                f.name, "Schema file", "WARNING",
                f"BIM Schema file '{SCHEMA_XLSX_NAME}' not found — cannot validate NWC."
            )
        result.build_granular_text()
        result.build_summary()
        if log_callback:
            log_callback(f"[NWC] Schema file not found: {SCHEMA_XLSX_NAME}")
        return result

    # Load schema
    if log_callback:
        log_callback(f"[NWC] Loading schema: {schema_path.name}")
    try:
        fields = _load_schema(schema_path)
    except Exception as exc:
        for f in nwc_files:
            result.files_checked.append(f.name)
            result.add_finding(f.name, "Schema load", "FAIL", f"Failed to load schema: {exc}")
        result.build_granular_text()
        result.build_summary()
        return result

    mandatory_count = sum(1 for f in fields if f.is_mandatory_for_phase(phase))
    if log_callback:
        log_callback(f"[NWC] Schema loaded: {len(fields)} fields, {mandatory_count} mandatory for '{phase}'")

    # Process each NWC file
    for nwc_file in nwc_files:
        result.files_checked.append(nwc_file.name)
        if log_callback:
            log_callback(f"[NWC] Opening: {nwc_file.name} (may take a moment)...")

        try:
            objects = _read_nwc(str(nwc_file))
        except Exception as exc:
            result.add_finding(
                nwc_file.name, "Open NWC", "FAIL",
                f"Failed to open via Navisworks COM: {exc}"
            )
            continue

        result.add_finding(
            nwc_file.name, "Object extraction", "PASS",
            f"Extracted {len(objects)} objects with BIM properties"
        )

        for obj in objects:
            _check_object(obj, fields, phase, result, nwc_file.name)

        if log_callback:
            fails = sum(1 for f in result.findings
                        if f.file_name == nwc_file.name and f.status == "FAIL")
            log_callback(f"[NWC] {nwc_file.name}: {fails} failure(s)")

    result.build_granular_text()
    result.build_summary()

    # Write outputs
    findings_path = output_folder / "nwc_findings.txt"
    findings_path.write_text(result.granular_text, encoding="utf-8")

    summary_path = output_folder / "nwc_summary.txt"
    summary_path.write_text(result.summary, encoding="utf-8")

    csv_path = output_folder / "nwc_validation_results.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["file", "check", "status", "details"])
        w.writeheader()
        for f in result.findings:
            w.writerow({"file": f.file_name, "check": f.check_name,
                         "status": f.status, "details": f.detail})

    if log_callback:
        log_callback(f"[NWC] Done — {'PASS' if result.overall_passed else 'FAIL'}")

    return result
