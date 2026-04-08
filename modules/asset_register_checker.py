"""
Asset Register Checker Module
=============================
Validates Sydney Metro asset register Excel files (.xlsx / .xlsm) against the
expected template structure (SM-22-00326845) and the BIM Configuration
Specification (SM-22-00340097).  Checks template structure, required sheets,
header alignment, file naming, and 30+ per-row field-presence / format /
cross-reference validations aligned with the IFC and NWC checkers.
"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules import ModuleResult, REFERENCE_DOCS

# ---------------------------------------------------------------------------
# Template constants (SM-22-00326845)
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = [
    "Standard governance",
    "Location Specification",
    "Location List - To Be Populated",
]

ASSET_LIST_SHEET = "Asset List - To Be Populated"
LOCATION_LIST_SHEET = "Location List - To Be Populated"

# Expected headers at row 6 in the Asset List sheet
REQUIRED_HEADER_CELLS = {
    "E6": "Asset Code",
    "F6": "Parent Asset Code",
    "G6": "Asset Type Code",
    "H6": "Asset Description",
    "I6": "Uniclass Title",
    "J6": "Uniclass Code",
    "T6": "GPS Coordinates",
    "AA6": "Coordinate Datum",
}

HEADER_ROW = 6
DATA_START_ROW = 7

# File-naming convention: SMWST*-*-*-*-AI-REG-*.xlsx
FILENAME_PATTERN = re.compile(
    r"^SM[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+-AI-REG-\d+.*\.xlsx$",
    re.I,
)

EMPTY_SENTINEL = {"", "N/A", "n/a", "NA", "na", "-", "--", "None", "none", "null", "TBD", "tbd"}

# ---------------------------------------------------------------------------
# All required column headers — aligned with BIM schema & IFC/NWC checkers
# (at least 30 distinct per-row checks)
# ---------------------------------------------------------------------------

# Columns mandatory for ALL asset rows (system and non-system)
MANDATORY_ALL_HEADERS = [
    # Identity & hierarchy
    "Asset Code",                       # 1
    "Parent Asset Code",                # 2
    "Asset Type Code",                  # 3
    "Asset Description",                # 4
    "Asset Label Description Hierarchy",# 5
    # Classification
    "Uniclass Code",                    # 6
    "Uniclass Title",                   # 7
    # IDs
    "TfNSW Asset ID",                   # 8
    "TfNSW Parent Asset ID",            # 9
    "Project Asset ID",                 # 10
    "Project Parent Asset ID",          # 11
    # Discipline
    "Discipline Code (Object)",         # 12
    "Discipline Description",           # 13
    "Sub-discipline Code (Object)",     # 14
    "Sub-discipline Description",       # 15
    # Location
    "Asset Location Code",              # 16
    "Asset Location Description Hierarchy",  # 17
    # Organisation
    "Asset Owner",                      # 18
    "Asset Maintainer Primary",         # 19
    "Asset Operator",                   # 20
    # Status
    "Asset Status Code",                # 21
    "Asset Status",                     # 22
    # Model traceability
    "Source Model",                     # 23
    "GUID",                             # 24
    "Maintenance Managed Item Flag",    # 25
    # Project
    "Project Contract Code",            # 26
    "Project Name",                     # 27
    # Asset type config
    "Asset Type Configuration Code (Specification)",  # 28
    # SMSA
    "SMSA ID",                          # 29
    "SMSA Name",                        # 30
]

# Additional columns mandatory only for non-system (physical) assets
MANDATORY_NON_SYSTEM_HEADERS = [
    "GPS Coordinates",                  # 31
    "Start Latitude",                   # 32
    "Start Longitude",                  # 33
    "End Latitude",                     # 34 (WARNING only — may be blank for point assets)
    "End Longitude",                    # 35 (WARNING only)
    "Start Km (Km)",                    # 36
    "End Km (Km)",                      # 37 (WARNING only)
    "Coordinate Datum",                 # 38
    "Local Government Area (LGA)",      # 39
]

# Headers where a WARNING is acceptable if blank (point assets etc.)
WARN_ONLY_IF_BLANK = {
    "End Latitude", "End Longitude", "End Km (Km)",
}

# Coordinate-datum expected value (case-insensitive)
EXPECTED_COORD_DATUM = "GDA2020/MGA zone 56"

# GPS format: expect decimal degrees pattern  e.g. -33.8688, 151.2093
GPS_PATTERN = re.compile(r"^-?\d{1,3}\.\d+$")

# Uniclass code format: e.g. Ss_25_10_30  (2-letter prefix, underscores, numbers)
UNICLASS_PATTERN = re.compile(r"^[A-Z][a-z]_\d{2}(_\d{2}){1,3}$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty(value: str) -> bool:
    return value in EMPTY_SENTINEL


def _is_system_asset(desc: str) -> bool:
    low = desc.lower()
    return "system" in low or "systems" in low


def _last_data_bounds(ws: Worksheet) -> Tuple[int, int]:
    last_row = 0
    last_col = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        row_has_data = False
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                row_has_data = True
                if cell.column > last_col:
                    last_col = cell.column
        if row_has_data:
            last_row = row[0].row
    return max(last_row, HEADER_ROW), max(last_col, 27)


def _find_header_column(ws: Worksheet, header: str, max_col: int) -> int | None:
    for col in range(1, max_col + 1):
        val = _normalize(ws.cell(row=HEADER_ROW, column=col).value)
        if val.lower() == header.lower():
            return col
    return None


def _export_csv(ws: Worksheet, path: Path, last_row: int, last_col: int) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1,
                                max_col=last_col, values_only=True):
            writer.writerow(["" if v is None else v for v in row])


# ---------------------------------------------------------------------------
# Validation  (structure → file naming → header alignment → row-level)
# ---------------------------------------------------------------------------

def _validate_workbook(xlsx_file: Path, output_dir: Path, result: ModuleResult) -> None:
    fname = xlsx_file.name

    # ── 0. File naming convention ──────────────────────────────────────
    if FILENAME_PATTERN.match(fname):
        result.add_finding(fname, "File naming convention", "PASS",
                           "Filename follows SM asset-register naming convention.")
    else:
        result.add_finding(fname, "File naming convention", "WARNING",
                           f"Filename '{fname}' may not follow the expected SM naming pattern "
                           "(e.g. SMWST*-*-*-*-AI-REG-*.xlsx).")

    # ── 1. Open workbook ───────────────────────────────────────────────
    try:
        wb = load_workbook(xlsx_file, data_only=True)
    except Exception as exc:
        result.add_finding(fname, "Open workbook", "FAIL", f"Cannot open file: {exc}")
        return

    sheet_names = wb.sheetnames

    # ── 2. Required sheets ─────────────────────────────────────────────
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in sheet_names]
    if missing_sheets:
        result.add_finding(fname, "Required sheets", "FAIL",
                           f"Missing sheet(s): {', '.join(missing_sheets)}")
    else:
        result.add_finding(fname, "Required sheets", "PASS",
                           "All required sheets present.")

    # ── 3. Asset List sheet existence ──────────────────────────────────
    ws_asset = wb[ASSET_LIST_SHEET] if ASSET_LIST_SHEET in sheet_names else None
    if ws_asset is None:
        result.add_finding(fname, "Asset list sheet", "FAIL",
                           f"Missing required sheet: {ASSET_LIST_SHEET}")
        wb.close()
        return

    # ── 4. Location List sheet has data ────────────────────────────────
    if LOCATION_LIST_SHEET in sheet_names:
        ws_loc = wb[LOCATION_LIST_SHEET]
        loc_has_data = any(
            ws_loc.cell(row=r, column=1).value is not None
            for r in range(2, min(ws_loc.max_row + 1, 10))
        )
        if loc_has_data:
            result.add_finding(fname, "Location List populated", "PASS",
                               "Location List sheet contains data.")
        else:
            result.add_finding(fname, "Location List populated", "WARNING",
                               "Location List sheet appears empty.")

    # ── 5. Header cell alignment ──────────────────────────────────────
    header_ok = True
    for cell_ref, expected in REQUIRED_HEADER_CELLS.items():
        actual = _normalize(ws_asset[cell_ref].value)
        if actual.lower() != expected.lower():
            header_ok = False
            result.add_finding(fname, "Header cell", "FAIL",
                               f"Cell {cell_ref}: expected '{expected}', "
                               f"found '{actual or '[blank]'}'")
    if header_ok:
        result.add_finding(fname, "Header cells", "PASS",
                           "All header cells match the Sydney Metro template.")

    # ── 6. Export CSV ─────────────────────────────────────────────────
    last_row, last_col = _last_data_bounds(ws_asset)
    csv_name = f"{xlsx_file.stem}_Asset_List.csv"
    csv_path = output_dir / csv_name
    try:
        _export_csv(ws_asset, csv_path, last_row, last_col)
        result.add_finding(fname, "CSV export", "PASS", f"Exported to {csv_name}")
    except Exception as exc:
        result.add_finding(fname, "CSV export", "WARNING", f"CSV export failed: {exc}")

    # ── 7. Resolve all column positions ───────────────────────────────
    all_headers = list(dict.fromkeys(
        MANDATORY_ALL_HEADERS + MANDATORY_NON_SYSTEM_HEADERS
    ))
    header_cols: dict[str, int] = {}
    missing_cols: list[str] = []
    for h in all_headers:
        col = _find_header_column(ws_asset, h, last_col)
        if col is not None:
            header_cols[h] = col
        else:
            missing_cols.append(h)

    if missing_cols:
        result.add_finding(fname, "Column headers present", "FAIL",
                           f"Missing column header(s): {', '.join(missing_cols[:15])}"
                           + (f" ...and {len(missing_cols)-15} more"
                              if len(missing_cols) > 15 else ""))
    else:
        result.add_finding(fname, "Column headers present", "PASS",
                           f"All {len(all_headers)} expected column headers found.")

    # We can still do row checks on whichever columns we did find
    found_mandatory_all = {h for h in MANDATORY_ALL_HEADERS if h in header_cols}
    found_non_system = {h for h in MANDATORY_NON_SYSTEM_HEADERS if h in header_cols}

    # ── 8. Row-level value checks ─────────────────────────────────────
    seen_codes: dict[str, int] = {}
    seen_asset_ids: dict[str, int] = {}
    seen_guids: dict[str, int] = {}
    rows_checked = 0
    fail_counts: dict[str, int] = {}  # track per-check failures for summary

    def _row_fail(check: str, detail: str) -> None:
        fail_counts[check] = fail_counts.get(check, 0) + 1
        # Only emit first 50 granular findings per check type to keep output manageable
        if fail_counts[check] <= 50:
            result.add_finding(fname, check, "FAIL", detail)

    def _row_warn(check: str, detail: str) -> None:
        result.add_finding(fname, check, "WARNING", detail)

    for row_num in range(DATA_START_ROW, last_row + 1):
        # Determine if the row has any data at all
        sample_vals = [
            _normalize(ws_asset.cell(row=row_num, column=c).value)
            for c in list(header_cols.values())[:8]
        ]
        if not any(sample_vals):
            continue
        rows_checked += 1

        desc_col = header_cols.get("Asset Description")
        desc = _normalize(ws_asset.cell(row=row_num, column=desc_col).value) if desc_col else ""
        is_system = _is_system_asset(desc)

        # --- Mandatory-for-all checks ---
        for h in found_mandatory_all:
            val = _normalize(ws_asset.cell(row=row_num, column=header_cols[h]).value)
            if _is_empty(val):
                _row_fail(h, f"Row {row_num}: missing {h}")

        # --- Non-system-only checks ---
        if not is_system:
            for h in found_non_system:
                val = _normalize(ws_asset.cell(row=row_num, column=header_cols[h]).value)
                if _is_empty(val):
                    if h in WARN_ONLY_IF_BLANK:
                        _row_warn(h, f"Row {row_num}: {h} blank (acceptable for point assets)")
                    else:
                        _row_fail(h, f"Row {row_num}: missing {h}")

        # --- Asset Code uniqueness ---
        code_col = header_cols.get("Asset Code")
        code = _normalize(ws_asset.cell(row=row_num, column=code_col).value) if code_col else ""
        if code and not _is_empty(code):
            if code in seen_codes:
                _row_fail("Duplicate Asset Code",
                          f"Row {row_num}: duplicate code '{code}' "
                          f"(first on row {seen_codes[code]})")
            else:
                seen_codes[code] = row_num

        # --- TfNSW Asset ID uniqueness ---
        id_col = header_cols.get("TfNSW Asset ID")
        asset_id = _normalize(ws_asset.cell(row=row_num, column=id_col).value) if id_col else ""
        if asset_id and not _is_empty(asset_id):
            if asset_id in seen_asset_ids:
                _row_fail("Duplicate TfNSW Asset ID",
                          f"Row {row_num}: duplicate ID '{asset_id}' "
                          f"(first on row {seen_asset_ids[asset_id]})")
            else:
                seen_asset_ids[asset_id] = row_num

        # --- GUID uniqueness ---
        guid_col = header_cols.get("GUID")
        guid = _normalize(ws_asset.cell(row=row_num, column=guid_col).value) if guid_col else ""
        if guid and not _is_empty(guid):
            if guid in seen_guids:
                _row_fail("Duplicate GUID",
                          f"Row {row_num}: duplicate GUID '{guid}' "
                          f"(first on row {seen_guids[guid]})")
            else:
                seen_guids[guid] = row_num

        # --- Coordinate Datum value check (case-insensitive) ---
        datum_col = header_cols.get("Coordinate Datum")
        if datum_col and not is_system:
            datum = _normalize(ws_asset.cell(row=row_num, column=datum_col).value)
            if datum and not _is_empty(datum):
                if datum.lower() != EXPECTED_COORD_DATUM.lower():
                    _row_fail("Coordinate Datum value",
                              f"Row {row_num}: expected '{EXPECTED_COORD_DATUM}', "
                              f"found '{datum}'")

        # --- GPS format validation ---
        gps_col = header_cols.get("GPS Coordinates")
        if gps_col and not is_system:
            gps = _normalize(ws_asset.cell(row=row_num, column=gps_col).value)
            if gps and not _is_empty(gps):
                # Accept comma-separated lat/lon pair
                parts = [p.strip() for p in gps.replace(";", ",").split(",")]
                if len(parts) < 2:
                    _row_warn("GPS format",
                              f"Row {row_num}: GPS value '{gps}' does not appear "
                              "to be a lat/lon pair")

        # --- Uniclass Code format ---
        uni_col = header_cols.get("Uniclass Code")
        if uni_col:
            uni = _normalize(ws_asset.cell(row=row_num, column=uni_col).value)
            if uni and not _is_empty(uni):
                if not UNICLASS_PATTERN.match(uni):
                    _row_warn("Uniclass Code format",
                              f"Row {row_num}: Uniclass Code '{uni}' may not "
                              "follow expected pattern (e.g. Ss_25_10_30)")

        # --- Asset Status Code valid values ---
        status_col = header_cols.get("Asset Status Code")
        if status_col:
            status_val = _normalize(ws_asset.cell(row=row_num, column=status_col).value)
            if status_val and not _is_empty(status_val):
                valid_codes = {
                    "ACTI", "INAC", "PLAN", "DCOM", "DISP", "active",
                    "inactive", "planned", "decommissioned", "disposed",
                }
                if status_val.upper() not in {v.upper() for v in valid_codes}:
                    _row_warn("Asset Status Code value",
                              f"Row {row_num}: unexpected status code '{status_val}'")

        # --- Maintenance Managed Item Flag (Y/N) ---
        mmi_col = header_cols.get("Maintenance Managed Item Flag")
        if mmi_col:
            mmi = _normalize(ws_asset.cell(row=row_num, column=mmi_col).value)
            if mmi and not _is_empty(mmi):
                if mmi.upper() not in ("Y", "N", "YES", "NO"):
                    _row_warn("MMI Flag value",
                              f"Row {row_num}: expected Y/N, found '{mmi}'")

        # --- Parent Asset Code must also exist as an Asset Code (cross-ref) ---
        parent_col = header_cols.get("Parent Asset Code")
        # Collected for post-loop check below

    # ── 9. Post-loop cross-reference: parent codes ────────────────────
    if header_cols.get("Parent Asset Code") and header_cols.get("Asset Code"):
        parent_col = header_cols["Parent Asset Code"]
        orphan_count = 0
        for row_num in range(DATA_START_ROW, last_row + 1):
            parent = _normalize(ws_asset.cell(row=row_num, column=parent_col).value)
            if parent and not _is_empty(parent) and parent not in seen_codes:
                orphan_count += 1
                if orphan_count <= 10:
                    result.add_finding(fname, "Parent Asset Code reference", "WARNING",
                                       f"Row {row_num}: Parent Asset Code '{parent}' "
                                       "not found in Asset Code column")
        if orphan_count > 10:
            result.add_finding(fname, "Parent Asset Code reference", "WARNING",
                               f"...and {orphan_count - 10} more orphan parent codes")

    # ── 10. Summary of truncated checks ───────────────────────────────
    for check, count in fail_counts.items():
        if count > 50:
            result.add_finding(fname, check, "FAIL",
                               f"...and {count - 50} additional '{check}' failure(s) "
                               f"({count} total)")

    # ── 11. Overall row-level verdict ─────────────────────────────────
    if rows_checked > 0:
        total_row_fails = sum(fail_counts.values())
        if total_row_fails == 0:
            result.add_finding(fname, "Value checks", "PASS",
                               f"All {rows_checked} data rows passed value checks.")
    else:
        result.add_finding(fname, "Value checks", "WARNING",
                           "No data rows found in the Asset List sheet.")

    wb.close()


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def run(input_folder: Path, output_folder: Path, log_callback=None) -> ModuleResult:
    """
    Scan *input_folder* for .xlsx / .xlsm files and validate each as an
    asset register.  Returns a ModuleResult with granular findings and summary.
    """
    result = ModuleResult(module_name="Asset Register")

    xlsx_files = sorted(
        list(input_folder.glob("*.xlsx")) + list(input_folder.glob("*.xlsm"))
    )
    # Filter out temp files
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$") and f.is_file()]

    if not xlsx_files:
        result.build_granular_text()
        result.build_summary()
        if log_callback:
            log_callback("[Asset Register] No .xlsx/.xlsm files found.")
        return result

    for xlsx_file in xlsx_files:
        result.files_checked.append(xlsx_file.name)
        if log_callback:
            log_callback(f"[Asset Register] Checking: {xlsx_file.name}")
        _validate_workbook(xlsx_file, output_folder, result)

    result.build_granular_text()
    result.build_summary()

    # Write granular findings file
    findings_path = output_folder / "asset_register_findings.txt"
    findings_path.write_text(result.granular_text, encoding="utf-8")

    # Write module summary file
    summary_path = output_folder / "asset_register_summary.txt"
    summary_path.write_text(result.summary, encoding="utf-8")

    # Write CSV
    csv_path = output_folder / "asset_register_validation_results.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["file", "check", "status", "details"])
        w.writeheader()
        for f in result.findings:
            w.writerow({"file": f.file_name, "check": f.check_name,
                         "status": f.status, "details": f.detail})

    if log_callback:
        log_callback(f"[Asset Register] Done — {'PASS' if result.overall_passed else 'FAIL'}")

    return result
